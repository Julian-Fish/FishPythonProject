'''
import maya.cmds as mc
from openpyxl import Workbook, load_workbook
#load excel
workRootDir = mc.workspace(q = 1, rootDirectory = 1)
wb = load_workbook(workRootDir + "test.xlsx")
ws = wb["kadai01"]
uuid = mc.ls("JAC", uuid=1)
#find number by uuid
for row in ws.values:
    if row[2] == uuid[0]:
        mc.rename("JAC", row[0])
'''
import PySide2
import maya.cmds as mc
from openpyxl import Workbook, load_workbook
import shutil
import maya.mel as mel
#load excel
workRootDir = mc.workspace(q = 1, rootDirectory = 1)
wb = load_workbook(workRootDir + "test.xlsx")
ws = wb.active

#create tableWidget
tableWidget = QTableWidget()
headerLabels = [u"学籍番号", u"名前", u"UUID"]
tableWidget.setColumnCount(len(headerLabels))
tableWidget.setRowCount(len(tuple(ws.rows)))
tableWidget.setHorizontalHeaderLabels(headerLabels)

tableWidget.verticalHeader().setVisible(False)

for row, rowData in enumerate(ws.values):
    if rowData[0] is None:
        break
    for col, v in enumerate(rowData):
        item = QTableWidgetItem(v)
        tableWidget.setItem(row, col, item)
        

tableWidget.show()
#find item
tableWidget.setItemSelected(item, 1)
=============================================================================
#newsheet = "kadai01"
#ws = wb.create_sheet(newsheet)
uuidcol = ws["c"]
i = 0
for row in ws.iter_rows(min_row = 1):
    if row[0].value is none:
        #file end
        #print "end"
        break
    num = row[0].value
    name = row[1].value
    filename = num + " " + name
    #create jac empty group
    jac = mc.group(em=true, name="jac")
    uuid = mc.ls("jac", uuid=1)
    uuidcol[i].value = uuid[0].encode("ascii", "ignore")
    i += 1
    #hiden jac empty group
    mc.setattr("jac.hiddeninoutliner", true)
    mc.outlinereditor("outlinerpanel1", edit = true, refresh = true)
    #lock jac, save and copy map file
    mc.locknode("jac", lock = 1)
    mel.eval("file -save")
    originmapname = "map1.mb"
    originmapdir = workrootdir + "scenes/" + originmapname
    targetmapdir = workrootdir + "scenes/" + filename + ".mb"
    shutil.copy(originmapdir, targetmapdir)
    #unlock and delete jac
    mc.locknode("jac", lock = 0)
    mc.delete(jac)
#save excel
wb.save(workrootdir + "test.xlsx")
=============================================================================
