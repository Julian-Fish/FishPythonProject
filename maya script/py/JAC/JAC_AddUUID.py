# -*- coding: utf-8 -*-
from PySide2 import QtWidgets
from PySide2.QtWidgets import *
from PySide2.QtGui import *
from PySide2.QtCore import *
from PySide2.QtUiTools import QUiLoader

from openpyxl import Workbook, load_workbook

from maya.app.general.mayaMixin import MayaQWidgetBaseMixin
import maya.cmds as mc
import maya.mel as mel
import maya.utils as mu
import shutil
import time

# 間接パスの指定
workRootDir = mc.workspace(q = 1, rootDirectory = 1)
targetDir = workRootDir + "scenes/"
print targetDir

class CopyWork(QObject):
    finished = Signal()
    progress = Signal(int)
    uuidList = Signal(list)

    def __init__(self, table, count, parent=None):
        super(CopyWork, self).__init__(parent)
        
        self.count = count
        self.table = table

    def copyMap(self):
        #generate authorStr and uuid
        _uuidList = []
        for i in range(0, self.count):
            uuid = mu.executeInMainThreadWithResult(self.doInMain, i)
            _uuidList.append(uuid[0])
            self.progress.emit(i + 1)
        self.uuidList.emit(_uuidList)
        self.finished.emit()
        
        
    def doInMain(self, i):
        number = self.table[i][0]
        name = self.table[i][1]
        authorStr = number + name
        
        # 1.add author to attribute
        # #create author node and set hidden
        # authorNode = mc.group(em = 1, name = "AuthorNode")
        # mc.setAttr("AuthorNode.hiddenInOutliner", True)
        # #add attr and lock it
        # mc.addAttr("AuthorNode", longName = "Author", dataType = "string")
        # mc.setAttr("AuthorNode.Author", authorStr, type = "string")
        # mc.setAttr("AuthorNode.Author", lock = 1)
        
        # 2.add author(number) to group name
        # create author node ,get uuid and  set hidden
        groupName = "Author_" + number
        authorNode = mc.group(em = 1, name = groupName)
        uuid = mc.ls(authorNode, uuid = 1)
        mc.setAttr(authorNode + ".hiddenInOutliner", True)
        
        #add origin uuid to attribute
        mc.addAttr(authorNode,longName = "signUUID", dataType = "string")
        mc.setAttr(authorNode + ".signUUID", uuid[0], type = "string")

        #lock node
        mc.lockNode(authorNode, lock = 1)
        #fresh outliner
        mc.outlinerEditor("outlinerPanel1", edit = 1, refresh = 1)
        
        #save and copy map
        mel.eval("file -save")
        #originMapName = "origin.mb"
        #originMap = workRootDir + "scenes/" + originMapName
        originMap = mc.file(q = 1, sn = 1)
        targetMapName = authorStr
        targetMap = targetDir + targetMapName + ".mb"
        print originMap
        print targetMap
        shutil.copy(originMap, targetMap)
        
        #unlock and delete node
        mc.lockNode(authorNode, lock = 0)
        mc.delete(authorNode)
        
        #uuid as result
        return uuid

## MainWindowを作るクラス
class MainWindow(MayaQWidgetBaseMixin, QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # # ウィンドウタイトルをUIから取得
        # self.setWindowTitle(self.UI.windowTitle())
        # # ウィジェットをセンターに配置
        # self.setCentralWidget(self.UI)
        
        # try to layout with code
        # create widget
        self.FileNameLabel = QLabel("Excel File:", self)
        self.FileNameValueLabel = QLabel("")
        Browse1Btn = QPushButton("Browse..")
        
        self.TargetDirLabel = QLabel("Directory:")
        self.TargetDirValueLabel = QLabel(targetDir + "scenes")
        self.Browse2Btn = QPushButton("Browse..")
        self.isDefaultDir = True

        self.CopyBtn = QPushButton("Copy Map && Generate UUID")
        self.LockBtn = QPushButton("Lock Node")
        self.UnlockBtn = QPushButton("Unlock Node")

        #excel file
        excelFileLabelHBox = QHBoxLayout()
        excelFileLabelHBox.addWidget(self.FileNameLabel)
        excelFileLabelHBox.addWidget(self.FileNameValueLabel)
        self.FileNameValueLabel.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.FileNameValueLabel.setFixedSize(200, 20)
        excelFileLabelHBox.addWidget(Browse1Btn)
        #-----------------------------------
        
        #target dir
        targetDirLabelHBox = QHBoxLayout()
        targetDirLabelHBox.addWidget(self.TargetDirLabel)
        targetDirLabelHBox.addWidget(self.TargetDirValueLabel)
        self.TargetDirValueLabel.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.TargetDirValueLabel.setFixedSize(200, 20)
        targetDirLabelHBox.addWidget(self.Browse2Btn)
        #-----------------------------------
        
        #excel view
        excelViewHBox = QHBoxLayout()
        #   sheet list
        sheetVBox = QVBoxLayout()
        self.sheetListWidget = QListWidget()
        self.sheetListWidget.setFixedSize(200, 300)
        #   create sheet button
        newSheetHBox = QHBoxLayout()
        self.newSheetLineEdit = QLineEdit()
        newSheetBtn = QPushButton("+")
        newSheetHBox.addWidget(self.newSheetLineEdit)
        newSheetHBox.addWidget(newSheetBtn)
        newSheetBtn.clicked.connect(self.createSheet)

        sheetVBox.addWidget(self.sheetListWidget)
        sheetVBox.addLayout(newSheetHBox)
        #sheetVBox.addWidget(newSheetBtn)
        self.sheetListWidget.currentItemChanged.connect(self.loadTable)
        #   table
        self.tableWidget = QTableWidget()
        headerLabels = [u"学籍番号", u"名前", u"UUID"]
        self.tableWidget.setColumnCount(len(headerLabels))
        self.tableWidget.setHorizontalHeaderLabels(headerLabels)

        excelViewHBox.addLayout(sheetVBox)
        excelViewHBox.addWidget(self.tableWidget)
        #-----------------------------------
        
        
        #lock/unlock button
        lockBtnHBox = QHBoxLayout()
        lockBtnHBox.addWidget(self.LockBtn)
        lockBtnHBox.addWidget(self.UnlockBtn)
        #-----------------------------------
        
        #whole
        layoutVBox = QVBoxLayout()
        layoutVBox.addLayout(excelFileLabelHBox)
        layoutVBox.addLayout(targetDirLabelHBox)
        layoutVBox.addLayout(excelViewHBox)
        layoutVBox.addWidget(self.CopyBtn)
        layoutVBox.addLayout(lockBtnHBox)
        #-----------------------------------
                                    
        self.setWindowTitle("JAC")
        
        widget = QWidget()
        widget.setLayout(layoutVBox)
        self.setCentralWidget(widget)
        
        # 接続
        Browse1Btn.clicked.connect(self.browseExcelFile)
        self.Browse2Btn.clicked.connect(self.browseTargetFolder)
        self.CopyBtn.clicked.connect(self.copyMap)
        self.LockBtn.clicked.connect(self.lockNode)
        self.UnlockBtn.clicked.connect(self.unlockNode)

        #for button disable test
        #self.CopyBtn.setEnabled(False)
        
    def browseTargetFolder(self):
        global targetDir
        #open dir dialog
        openDir = QFileDialog.getExistingDirectory(self, "Target Directory", "", QFileDialog.ShowDirsOnly)
        if openDir != "":
            #fix dir
            if openDir[-1] != "/":
                openDir += "/"

            self.TargetDirValueLabel.setText(openDir)
            print openDir
            targetDir = openDir

    def browseExcelFile(self):
        #open file dialog
        fileDir = QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files(*.xlsx)")
        self.currentExcelFileDir = fileDir[0]

        splitStr = fileDir[0].split("/")
        fileName = splitStr[-1]
        #print str[-1]
        # self.UI.FileNameValue.setText(fileName)
        if fileName != "":
            self.FileNameValueLabel.setText(fileName)
            #load file
            self.wb = load_workbook(fileDir[0])
            self.loadSheet()

    def loadSheet(self):
        self.tableWidget
        
        self.sheetListWidget.clear()
        sheetList = self.wb.sheetnames

        for v in sheetList[1::]:
            item = QListWidgetItem(v)
            self.sheetListWidget.addItem(item)

    def loadTable(self):
        self.tableWidget.clearContents()

        sheetName = self.sheetListWidget.currentItem().text()
        print sheetName

        ws = self.wb[sheetName]
        rowCount = len(tuple(ws.rows))
        self.tableWidget.setRowCount(rowCount)

        for row, rowData in enumerate(ws.values):
            if rowData[0] is None:
                break
            for col, data in enumerate(rowData):
                item = QTableWidgetItem(data)
                self.tableWidget.setItem(row, col, item)
    
    def createSheet(self):
        newSheetName = self.newSheetLineEdit.text()
        source = self.wb[self.wb.sheetnames[0]]
        ws = self.wb.copy_worksheet(source)
        ws.title = newSheetName
        self.save()

        item = QListWidgetItem(newSheetName)
        self.sheetListWidget.addItem(item)
        self.sheetListWidget.setCurrentItem(item)
 
    def copyMap(self):
        ws = self.wb[self.sheetListWidget.currentItem().text()]
        self.count = 0
        for row in ws.values:
            if row[0] is None:
                #file end
                break
            self.count += 1
        
        self.ProgressDialog = QProgressDialog("Copying Map", "Close", 0, self.count)
        self.ProgressDialog.setWindowTitle("Copy Progress")
        self.ProgressDialog.show()
        table = tuple(ws.values)
        
        #set operation thread
        self.thread = QThread()
        self.copyWork = CopyWork(table, self.count)
        #disable button to avoid double click(crush issue)
        self.CopyBtn.setEnabled(False)
        self.copyWork.copyMap()  # run without thread for test
        
        self.copyWork.moveToThread(self.thread)
        self.copyWork.uuidList.connect(self.emitUuidList)
        self.copyWork.progress.connect(self.emitProgress)
        self.copyWork.finished.connect(self.copyDone)
        self.copyWork.finished.connect(self.thread.quit)
        self.copyWork.finished.connect(self.copyWork.deleteLater)
        #enable button
        self.copyWork.finished.connect(
            lambda: self.CopyBtn.setEnabled(True)
        )

        self.thread.started.connect(self.copyWork.copyMap)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.start()
    
    def emitUuidList(self, uuidList):
        #print uuidList
        currentSheet = self.sheetListWidget.currentItem().text()

        #input uuid to col 3
        ws = self.wb[currentSheet]
        uuidCol = tuple(ws["C"])

        for i in range(0, len(uuidList)):
            uuidCol[i].value = uuidList[i]

        self.save()
        self.loadTable()

    def emitProgress(self, p):
        self.copiedCountStr = str(p)
        label = self.copiedCountStr + "/" + str(self.count)
        self.ProgressDialog.setLabelText("Copying Map: " + label)
        self.ProgressDialog.setValue(p)
    
    def copyDone(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Done")
        msgBox.setText(r"<b>Copy Done<b>")
        msgBox.setTextFormat(Qt.RichText)
        msgBox.setInformativeText(self.copiedCountStr + " Map Copied.")
        #msgBox.setDetailedText(self.copyCount + " Map Copied.")
        msgBox.addButton(QMessageBox.Ok)
        
        msgBox.exec_()
        
    def lockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 1)
            print node + " locked"
    def unlockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 0)
            print node + " unlocked"

    def save(self):
        self.wb.save(self.currentExcelFileDir)

## MainWindowの起動
def main():
    window = MainWindow()
    window.show()

if __name__ == '__main__':

    main()