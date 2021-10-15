# -*- coding: utf-8 -*-
from PySide2 import QtWidgets
from PySide2.QtUiTools import QUiLoader
from openpyxl import Workbook, load_workbook
from maya.app.general.mayaMixin import MayaQWidgetBaseMixin
import maya.cmds as mc
import shutil
import maya.mel as mel

# 間接パスの指定
UIFILEPATH = "D:/19cu0217/FishPythonProject/maya script/qtui/JAC.ui"

## MainWindowを作るクラス
class MainWindow(MayaQWidgetBaseMixin, QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # UIのパスを指定
        self.UI = QUiLoader().load(UIFILEPATH)
        # ウィンドウタイトルをUIから取得
        self.setWindowTitle(self.UI.windowTitle())
        # ウィジェットをセンターに配置
        self.setCentralWidget(self.UI)
        
        # 接続
        self.UI.BrowseBtn.clicked.connect(self.openFileDialog)
        self.UI.SheetNameList.itemClicked.connect(self.tableLoad)
        self.UI.NewSheetBtn.clicked.connect(self.newSheet)
        self.UI.GenerateBtn.clicked.connect(self.generateMap_UUID)

    def openFileDialog(self):
        #open file dialog
        fileDir = QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files(*.xlsx)")
        splitStr = fileDir[0].split("/")
        fileName = splitStr[-1]
        #print str[-1]
        self.UI.FileNameValue.setText(fileName)
        #load file
        self.wb = load_workbook(fileDir[0])
        
        #fill sheet list
        sheetList = self.wb.sheetnames
        for v in sheetList[1::]:
            item = QtWidgets.QListWidgetItem(v)
            self.UI.SheetNameList.addItem(item)

        
    def tableLoad(self):
        #table load
        tableWidget = self.UI.AuthorTable
        selectItems = self.UI.SheetNameList.selectedItems()
        sheetName = selectItems[0].text()
        ws = self.wb[sheetName]
        
        rowCount = len(tuple(ws.rows))
        tableWidget.setRowCount(rowCount)
        headerLabels = [u"学籍番号", u"名前", u"UUID"]
        tableWidget.setColumnCount(len(headerLabels))
        tableWidget.setHorizontalHeaderLabels(headerLabels)
        
        for row, rowData in enumerate(ws.values):
            if rowData[0] is None:
                break
            for col, data in enumerate(rowData):
                item = QtWidgets.QTableWidgetItem(data)
                tableWidget.setItem(row, col, item)
        tableWidget.show()
    
    def newSheet(self):
        print "new sheet"
    
    def generateMap_UUID(self):
        print "generateMap_UUID"
        
        
## MainWindowの起動
def main():
    window = MainWindow()
    window.show()

if __name__ == '__main__':
    main()