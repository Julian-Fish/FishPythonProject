# -*- coding: utf-8 -*-
from PySide2 import QtWidgets
from PySide2.QtWidgets import *
from PySide2.QtGui import *
from PySide2.QtUiTools import QUiLoader

from openpyxl import Workbook, load_workbook

from maya.app.general.mayaMixin import MayaQWidgetBaseMixin
import maya.cmds as mc
import shutil
import maya.mel as mel

# 間接パスの指定
UIFILEPATH = "D:/19cu0217/FishPythonProject/maya script/qtui/JAC_lockNode.ui"
workRootDir = mc.workspace(q = 1, rootDirectory = 1)

print workRootDir

class CopyOperation(QObject):
    count = Signal(int)
    
    def copyMap(self, number, name):
        

## MainWindowを作るクラス
class MainWindow(MayaQWidgetBaseMixin, QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # # UIのパスを指定
        # self.UI = QUiLoader().load(UIFILEPATH)
        # # ウィンドウタイトルをUIから取得
        # self.setWindowTitle(self.UI.windowTitle())
        # # ウィジェットをセンターに配置
        # self.setCentralWidget(self.UI)
        
        # try to layout with code
        # create widget
        self.FileNameLabel = QLabel("Excel File:", self)
        self.FileNameValueLabel = QLabel("")
        self.BrowseBtn = QPushButton("Brose..")
        self.CopyBtn = QPushButton("Copy Map")
        self.LockBtn = QPushButton("Lock Node")
        self.UnlockBtn = QPushButton("Unlock Node")

        # layout
        layoutHBox1 = QHBoxLayout()
        layoutHBox1.addWidget(self.FileNameLabel)
        layoutHBox1.addWidget(self.FileNameValueLabel)
        layoutHBox1.addWidget(self.BrowseBtn)
        
        layoutHBox2 = QHBoxLayout()
        layoutHBox2.addWidget(self.LockBtn)
        layoutHBox2.addWidget(self.UnlockBtn)
        
        layoutVBox = QVBoxLayout()
        layoutVBox.addLayout(layoutHBox1)
        layoutVBox.addWidget(self.CopyBtn)
        layoutVBox.addLayout(layoutHBox2)
                                    
        # self.setLayout(layoutVBox)#
        self.setWindowTitle("JAC")
        
        widget = QWidget()
        widget.setLayout(layoutVBox)
        self.setCentralWidget(widget)
        
        # 接続
        self.BrowseBtn.clicked.connect(self.openFileDialog)
        self.CopyBtn.clicked.connect(self.copyMap)
        self.LockBtn.clicked.connect(self.lockNode)
        self.UnlockBtn.clicked.connect(self.unlockNode)
        # self.UI.BrowseBtn.clicked.connect(self.openFileDialog)
        # self.UI.CopyBtn.clicked.connect(self.copyMap)
        
    def openFileDialog(self):
        #open file dialog
        fileDir = QtWidgets.QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files(*.xlsx)")
        splitStr = fileDir[0].split("/")
        fileName = splitStr[-1]
        #print str[-1]
        # self.UI.FileNameValue.setText(fileName)
        self.FileNameValueLabel.setText(fileName)
        #load file
        self.wb = load_workbook(fileDir[0])
    
    def copyMap(self):
        ws = self.wb[self.wb.sheetnames[0]]
        count = 0
        for row in ws.values:
            if row[0] is None:
                #file end
                break
            count += 1
        
        self.ProgressDialog = QProgressDialog("Copying Map", "Cancel", 0, count)
        self.ProgressDialog.setWindowTitle("Copy Progress")
        table = tuple(ws.values)
        
        for i in range(0, count):
            if self.ProgressDialog.wasCanceled():
                #copy canceled
                break
            
            #generate authorStr
            authorStr = table[i][0] + table[i][1]
            
            # 1.add author to attribute
            # #create author node and set hidden
            # authorNode = mc.group(em = 1, name = "AuthorNode")
            # mc.setAttr("AuthorNode.hiddenInOutliner", True)
            # #add attr and lock it
            # mc.addAttr("AuthorNode", longName = "Author", dataType = "string")
            # mc.setAttr("AuthorNode.Author", authorStr, type = "string")
            # mc.setAttr("AuthorNode.Author", lock = 1)
            
            # 2.add author(number) to group name
            # create author node and set hidden
            groupName = "Author_" + table[i][0]
            print groupName
            authorNode = mc.group(em = 1, name = groupName)
            mc.setAttr(groupName + ".hiddenInOutliner", True)
            
            #lock node
            mc.lockNode(authorNode, lock = 1)
            #fresh outliner
            mc.outlinerEditor("outlinerPanel1", edit = 1, refresh = 1)
            
            #save and copy map
            mel.eval("file -save")
            originMapName = "origin.mb"
            targetMapName = authorStr
            originMap = workRootDir + "scenes/" + originMapName
            targetMap = workRootDir + "scenes/" + targetMapName + ".mb"
            shutil.copy(originMap, targetMap)
            
            #unlock and delete node
            mc.lockNode(authorNode, lock = 0)
            mc.delete(authorNode)
            count += 1
            
            # set progress
            self.ProgressDialog.setValue(count)
            
        self.copyCount = str(count)
        self.copyDone()
    
    def createAuthorNode:
        
    
    def copyDone(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Done")
        msgBox.setText(r"<b>Copy Done<b>")
        msgBox.setTextFormat(QtGui.Qt.RichText)
        msgBox.setInformativeText(self.copyCount + " Map Copied.")
        #msgBox.setDetailedText(self.copyCount + " Map Copied.")
        msgBox.addButton(QMessageBox.Ok)
        
        msgBox.exec_()
        
    def lockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 1)
            print node + " locked"
    def unlockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 0)
            print node + " unlocked"
## MainWindowの起動
def main():
    window = MainWindow()
    window.show()

if __name__ == '__main__':
    main()