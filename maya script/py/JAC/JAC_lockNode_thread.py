# -*- coding: utf-8 -*-
from PySide2 import QtWidgets
from PySide2.QtWidgets import *
from PySide2.QtGui import *
from PySide2.QtCore import *
from PySide2.QtUiTools import QUiLoader

from openpyxl import Workbook, load_workbook

from maya.app.general.mayaMixin import MayaQWidgetBaseMixin
import maya.cmds as mc
import maya.mel as mel
import maya.utils as mu
import shutil
import time

# 間接パスの指定
workRootDir = mc.workspace(q = 1, rootDirectory = 1)
targetDir = workRootDir
print targetDir

class CopyWork(QObject):
    finished = Signal()
    progress = Signal(int)
    def __init__(self, table, count, parent=None):
        super(CopyWork, self).__init__(parent)
        
        self.count = count
        self.table = table

    def copyMap(self):
        #generate authorStr
        for i in range(0, self.count):
            mu.executeInMainThreadWithResult(self.doInMain, i)
            self.progress.emit(i + 1)
        self.finished.emit()
        
    def doInMain(self, i):
        number = self.table[i][0]
        name = self.table[i][1]
        authorStr = number + name
        
        # 1.add author to attribute
        # #create author node and set hidden
        # authorNode = mc.group(em = 1, name = "AuthorNode")
        # mc.setAttr("AuthorNode.hiddenInOutliner", True)
        # #add attr and lock it
        # mc.addAttr("AuthorNode", longName = "Author", dataType = "string")
        # mc.setAttr("AuthorNode.Author", authorStr, type = "string")
        # mc.setAttr("AuthorNode.Author", lock = 1)
        
        # 2.add author(number) to group name
        # create author node and set hidden
        groupName = "Author_" + number
        authorNode = mc.group(em = 1, name = groupName)
        mc.setAttr(groupName + ".hiddenInOutliner", True)
        
        #lock node
        mc.lockNode(authorNode, lock = 1)
        #fresh outliner
        mc.outlinerEditor("outlinerPanel1", edit = 1, refresh = 1)
        
        #save and copy map
        mel.eval("file -save")
        #originMapName = "origin.mb"
        #originMap = workRootDir + "scenes/" + originMapName
        originMap = mc.file(q = 1, sn = 1)
        targetMapName = authorStr
        targetMap = targetDir + targetMapName + ".mb"
        print originMap
        print targetMap
        shutil.copy(originMap, targetMap)
        
        #unlock and delete node
        mc.lockNode(authorNode, lock = 0)
        mc.delete(authorNode)
        
        return groupName

## MainWindowを作るクラス
class MainWindow(MayaQWidgetBaseMixin, QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # # ウィンドウタイトルをUIから取得
        # self.setWindowTitle(self.UI.windowTitle())
        # # ウィジェットをセンターに配置
        # self.setCentralWidget(self.UI)
        
        # try to layout with code
        # create widget
        self.FileNameLabel = QLabel("Excel File:", self)
        self.FileNameValueLabel = QLabel("")
        self.Browse1Btn = QPushButton("Browse..")
        
        self.TargetDirLabel = QLabel("Directory:")
        self.TargetDirValueLabel = QLabel(targetDir + "scenes")
        self.Browse2Btn = QPushButton("Browse..")
        self.isDefaultDir = True

        self.CopyBtn = QPushButton("Copy Map")
        self.LockBtn = QPushButton("Lock Node")
        self.UnlockBtn = QPushButton("Unlock Node")

        # layout
        layoutHBox1 = QHBoxLayout()
        layoutHBox1.addWidget(self.FileNameLabel)
        layoutHBox1.addWidget(self.FileNameValueLabel)
        self.FileNameValueLabel.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.FileNameValueLabel.setFixedSize(200, 20)
        layoutHBox1.addWidget(self.Browse1Btn)
        
        layoutHBox2 = QHBoxLayout()
        layoutHBox2.addWidget(self.TargetDirLabel)
        layoutHBox2.addWidget(self.TargetDirValueLabel)
        self.TargetDirValueLabel.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.TargetDirValueLabel.setFixedSize(200, 20)
        layoutHBox2.addWidget(self.Browse2Btn)

        layoutHBox3 = QHBoxLayout()
        layoutHBox3.addWidget(self.LockBtn)
        layoutHBox3.addWidget(self.UnlockBtn)
        
        layoutVBox = QVBoxLayout()
        layoutVBox.addLayout(layoutHBox1)
        layoutVBox.addLayout(layoutHBox2)
        layoutVBox.addWidget(self.CopyBtn)
        layoutVBox.addLayout(layoutHBox3)
                                    
        # self.setLayout(layoutVBox)#
        self.setWindowTitle("JAC")
        
        widget = QWidget()
        widget.setLayout(layoutVBox)
        self.setCentralWidget(widget)
        
        # 接続
        self.Browse1Btn.clicked.connect(self.browseExcelFile)
        self.Browse2Btn.clicked.connect(self.browseTargetFolder)
        self.CopyBtn.clicked.connect(self.copyMap)
        self.LockBtn.clicked.connect(self.lockNode)
        self.UnlockBtn.clicked.connect(self.unlockNode)
        
    def browseTargetFolder(self):
        global targetDir
        #open dir dialog
        openDir = QFileDialog.getExistingDirectory(self, "Target Directory", "", QFileDialog.ShowDirsOnly)
        if openDir != "":
            #fix dir
            if openDir[-1] != "/":
                openDir += "/"

            self.TargetDirValueLabel.setText(openDir)
            print openDir
            targetDir = openDir

    def browseExcelFile(self):
        #open file dialog
        fileDir = QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files(*.xlsx)")
        splitStr = fileDir[0].split("/")
        fileName = splitStr[-1]
        #print str[-1]
        # self.UI.FileNameValue.setText(fileName)
        self.FileNameValueLabel.setText(fileName)
        #load file
        self.wb = load_workbook(fileDir[0])
    
    def copyMap(self):
        ws = self.wb[self.wb.sheetnames[0]]
        self.count = 0
        for row in ws.values:
            if row[0] is None:
                #file end
                break
            self.count += 1
        
        self.ProgressDialog = QProgressDialog("Copying Map", "Close", 0, self.count)
        self.ProgressDialog.setWindowTitle("Copy Progress")
        self.ProgressDialog.show()
        table = tuple(ws.values)
        
        #set operation thread
        self.thread = QThread()
        self.copyWork = CopyWork(table, self.count)
        #self.copyWork.copyMap()  # run without thread for test
        
        self.copyWork.moveToThread(self.thread)
        self.copyWork.progress.connect(self.emitProgress)
        self.copyWork.finished.connect(self.copyDone)
        self.copyWork.finished.connect(self.thread.quit)
        self.copyWork.finished.connect(self.copyWork.deleteLater)

        self.thread.started.connect(self.copyWork.copyMap)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.start()
    
    def emitProgress(self, p):
        self.copiedCountStr = str(p)
        label = self.copiedCountStr + "/" + str(self.count)
        self.ProgressDialog.setLabelText("Copying Map: " + label)
        self.ProgressDialog.setValue(p)
    
    def copyDone(self):
        print "copy done"
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Done")
        msgBox.setText(r"<b>Copy Done<b>")
        msgBox.setTextFormat(Qt.RichText)
        msgBox.setInformativeText(self.copiedCountStr + " Map Copied.")
        #msgBox.setDetailedText(self.copyCount + " Map Copied.")
        msgBox.addButton(QMessageBox.Ok)
        
        msgBox.exec_()
        
    def lockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 1)
            print node + " locked"
    def unlockNode(self):
        nodes = mc.ls(sl = 1)
        for node in nodes:
            mc.lockNode(node, lock = 0)
            print node + " unlocked"
## MainWindowの起動
def main():
    window = MainWindow()
    window.show()

if __name__ == '__main__':

    main()