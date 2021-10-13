# -*- coding: utf-8 -*-
from PySide2 import QtWidgets
from PySide2.QtWidgets import *
from PySide2 import QtGui
from PySide2.QtUiTools import QUiLoader
from openpyxl import Workbook, load_workbook
from maya.app.general.mayaMixin import MayaQWidgetBaseMixin
import maya.cmds as mc
import shutil
import maya.mel as mel

# 間接パスの指定
UIFILEPATH = "D:/19cu0217/maya/FishPythonProject/maya script/qtui/JAC_lockNode.ui"
workRootDir = mc.workspace(q = 1, rootDirectory = 1)

print workRootDir

## MainWindowを作るクラス
class MainWindow(MayaQWidgetBaseMixin, QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # # UIのパスを指定
        # self.UI = QUiLoader().load(UIFILEPATH)
        # # ウィンドウタイトルをUIから取得
        # self.setWindowTitle(self.UI.windowTitle())
        # # ウィジェットをセンターに配置
        # self.setCentralWidget(self.UI)
        
        # try to layout with code
        # create widget
        FileNameLabel = QLabel("FileName:", self)
        self.FileNameValueLabel = QLabel("")
        BrowseBtn = QPushButton("Brose..")
        CopyBtn = QPushButton("Copy Map")

        # layout
        layoutHBox = QHBoxLayout(self)
        
        layoutHBox.addWidget(FileNameLabel)
        layoutHBox.addWidget(self.FileNameValueLabel)
        layoutHBox.addWidget(BrowseBtn)
        
        layoutVBox = QVBoxLayout(self)
        layoutVBox.addLayout(layoutHBox)
        layoutVBox.addWidget(CopyBtn)
                                    
        # self.setLayout(layoutVBox)#
        self.setWindowTitle("JAC")
        
        widget = QWidget(self)
        widget.setLayout(layoutVBox)
        self.setCentralWidget(widget)
        
        # 接続
        BrowseBtn.clicked.connect(self.openFileDialog)
        CopyBtn.clicked.connect(self.copyMap)
        # self.UI.BrowseBtn.clicked.connect(self.openFileDialog)
        # self.UI.CopyBtn.clicked.connect(self.copyMap)
        
    def openFileDialog(self):
        #open file dialog
        fileDir = QtWidgets.QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files(*.xlsx)")
        splitStr = fileDir[0].split("/")
        fileName = splitStr[-1]
        #print str[-1]
        # self.UI.FileNameValue.setText(fileName)
        self.FileNameValueLabel.setText(fileName)
        #load file
        self.wb = load_workbook(fileDir[0])
    
    def copyMap(self):
        ws = self.wb[self.wb.sheetnames[0]]
        count = 0
        for row in ws.values:
            if row[0] is None:
                #file end
                break
            #generate authorStr
            authorStr = row[0] + row[1]
            
            #create author node and set hidden
            authorNode = mc.group(em = 1, name = "AuthorNode")
            mc.setAttr("AuthorNode.hiddenInOutliner", True)
            #add attr and lock it
            mc.addAttr("AuthorNode", longName = "Author", dataType = "string")
            mc.setAttr("AuthorNode.Author", authorStr, type = "string")
            mc.setAttr("AuthorNode.Author", lock = 1)
            #lock node
            mc.lockNode(authorNode, lock = 1)
            #fresh outliner
            mc.outlinerEditor("outlinerPanel1", edit = 1, refresh = 1)
            
            #save and copy map
            mel.eval("file -save")
            originMapName = "origin.mb"
            targetMapName = authorStr
            originMap = workRootDir + "scenes/" + originMapName
            targetMap = workRootDir + "scenes/" + targetMapName + ".mb"
            shutil.copy(originMap, targetMap)
            
            #unlock and delete node
            mc.lockNode(authorNode, lock = 0)
            mc.delete(authorNode)
            count += 1
            
        self.copyCount = count   
        self.copyDone()
        
    def copyDone(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Done")
        msgBox.setText(r"<b>Copy Done<b>")
        msgBox.setTextFormat(QtGui.Qt.RichText)
        msgBox.setInformativeText(self.copyCount + " Map Copied.")
        #msgBox.setDetailedText(self.copyCount + " Map Copied.")
        msgBox.addButton(QMessageBox.Ok)
        
        msgBox.exec_()
        
## MainWindowの起動
def main():
    window = MainWindow()
    window.show()

if __name__ == '__main__':
    main()